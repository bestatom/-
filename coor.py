# -*- coding: UTF-8 -*-  
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
wb = load_workbook('10086.xlsx')
#print wb.get_sheet_names()
def fill_excel(ip,local):
	ip = str(ip)
	ws3 = wb[ip]
	ws3[local] = 1
	fi = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00',fill_type='solid')
	ws3[local].fill = fi
	ws3[local].fill = fi
	wb.save('10086.xlsx')



def coordubate(list1):
	list5 = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W']
	list6 = ['21','22','23','25','53','69','135','136','137','138','139','161','445','1433','1521','3306','3389','5631','5632','5900','6379','8080']
	list2 = []
	list3 = []
	list7 = []
	list8 = []
	for index,item in enumerate(list1):
		if len(item)>6:
			list2.append(index)
		else:
			list3.append(index)
	for index,item in enumerate(list2):
		if index == 0:
			pass
		else:
			list7.append(list1[list2[index-1]:list2[index]])
	list7.append(list1[list2[-1]:])

	for i in list7:
		list4 = []
		a = i[0].split('.')
		list4.append(int(a[3])+3)
		for index,item in enumerate(list6):
			for x in i[1:]:
				if x == item:
					list4.append(list5[index])
		for i in list4[1:]:
			list8.append(i+str(list4[0]))
	return list8



if __name__ == '__main__':
	list9 = ['218.203.101.2', '22','23' ,'218.203.101.6', '22', '218.203.101.9', '22', '218.203.101.13', '22', '218.203.101.28', '53', '218.203.101.29', '53', '218.203.101.36', '53']
	local = coordubate(list9)
	#print local
	#for i in local:
		#i = str(i)
		#fill_excel('218.203.101',i)
	print 'down'

import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from coor import coordubate
list5 = []
list5 = os.listdir('/root/Desktop/ip')
def fun(filename):
	list1 = []
	list2 = []
	list3 = []
	list4 = []
	for line in open(filename):
		list1.append(line)
	for i in list1:
		if 'open' in i or '218.203' in i or '211.138' in i :
			list2.append(i)
	for index,value in enumerate(list2):
		if 'open' in value :
			if '218.203' in list2[index-1][21:37] or '211.138' in list2[index-1][21:37]:
				list3.append(list2[index-1][21:37])
			list3.append(list2[index][0:4])
	for i in list3:
		i = i.replace('\r','').replace('\n','').replace('/t','').replace('/','')
		if i != '':
			list4.append(i)
	return list4

def fill_excel(ip,local):
	wb = load_workbook('10086.xlsx')
	ip = str(ip)
	ws3 = wb[ip]
	ws3[local] = 1
	fi = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00',fill_type='solid')
	ws3[local].fill = fi
	ws3[local].fill = fi
	wb.save('10086.xlsx')


if __name__ == '__main__':
	for i in list5:
		if 'py' not in i:
			list7 =  fun(i)
			print list7
			i = i.replace('..txt','').replace('.txt','')
			if list7 == []:
				pass
			else:
				local = coordubate(list7)
				print local
				for lo in local:
					lo = str(lo)
					print lo
					fill_excel(i,lo)
		print 1
